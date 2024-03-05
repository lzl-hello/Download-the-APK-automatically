import openpyxl
import re
import requests
import os.path
from sys import argv
from openpyxl import Workbook
from datetime import datetime
import threading
from concurrent.futures import ThreadPoolExecutor


url = ''
count = 0
lock = threading.Lock()  # 用于线程同步


def download_apk(num, url, name, version):
    """
    Args:
        url: apk下载链接
        num: 索引
        name: apk名称
        version: apk版本
    """
    global count
    succeed = 'Succeed'
    failure = 'Failure'

    print('第' + str(num-2) + '条url:\n' + url)

    # 笑死，第一个response获取到的不知道是些啥，加了些参数才能拿到
    # response = requests.get(url, stream=True)
    response = requests.get(url,headers={'User-Agent': 'Mozilla/5.0'}, timeout=10, stream=True)

    # 从响应头中获取包名
    if 'Content-Disposition' in response.headers:
        content_disposition = response.headers['Content-Disposition']
        packagename = re.findall('filename=(.+)', content_disposition)[0]
    else:
        packagename = os.path.basename(url)
    # 把参数部分去掉
    packagename = packagename.split('?')[0]

    # 下载的apk命名，多余了家人们，人家的filename已经名字版本都有了，就多了个apkpure，先这样吧 但是直接用它的filename中文会乱码
    filename = ""
    if name:
        filename += name + '_'
    else:
        filename += 'unknown_'
    if version:
        filename += version
    else:
        filename += 'unknown'
    filename += '.apk'

    # TODO: 获取文件大小和类型----解决了

    if 'Content-Length' in response.headers:
        # print(response.headers['Content-Length'])
        filesize = round(float(response.headers['Content-Length']) / 1048576, 2)
    else:
        filesize = 0  # 或者其他默认值

    # 判断文件类型和大小
    apk_format = 'application/vnd.android.package-archive'
    if response.headers['Content-Type'] == apk_format and filesize < 100.00:
    # if filesize < 100.00:
        print('文件类型：' + response.headers['Content-Type'] + "\n" +
              '文件大小：' + str(filesize) + 'M' + "\n" +
              '文件名：' + str(filename) + "\n" +
              'packagename：' + str(packagename))

        # 下载文件
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36 Edg/122.0.0.0',
            'Connection': 'keep-alive', }

        file = requests.get(url, headers=headers, timeout=10)
        # 区别不同类别的apk
        folder_path = '工具apk/'
        with open(folder_path + filename, 'wb') as apk:
            apk.write(file.content)
            print(succeed + "\n")
            # 使用锁来保护count变量的更新
            with lock:
                count += 1

            # count += 1
        # 返回内容
        dicts = [url, succeed, filename]
        return dicts
    else:
        print('文件类型:' + response.headers['Content-Type'] + "\n" +
              '文件大小:' + str(filesize) + 'M' + "\n" +
              failure + "\n")
        dicts = [url, failure, failure]
        return dicts



def deal_excel(path):
    """
    从Excel中获取下载链接
    :param path: 表格的绝对路径
    """

    global url
    start = datetime.now()
    wb = openpyxl.load_workbook(path)
    table = wb.get_sheet_by_name('Result 1')
    # 删除空行并计算有数据的行数
    max_row = 0

    for row in reversed(list(table.iter_rows())):
        # 如果第二列（URL 列）为空，则删除该行
        if row[1].value is None:
            table.delete_rows(row[0].row, 1)
        else:
            max_row += 1

    print('共' + str(max_row-2) + '条url.\n')

    wb2 = Workbook()
    ws = wb2.active

    # 前两行没用，跳过
    max_threads = 3
    with ThreadPoolExecutor(max_workers=max_threads) as executor:
        for row in range(3, max_row+1):
            name = table.cell(row=row, column=1).value
            url = table.cell(row=row, column=2).value
            version = table.cell(row=row, column=3).value

            # 检查http协议头
            if not re.match(r'^https?:/{2}\w.+$', url):
                url = "http://" + url

            # 提交任务给线程池
            executor.submit(download_apk, row, url, name, version)

        # 单线程 就每个循环调用一次download_apk
        # dicts = download_apk(row, url, name, version)

        # 新建Excel表，记录原始地址、下载结果和文件名，算了，保存了，直接看下载结果吧
        # ws['A' + str(row)] = dicts[0]
        # ws['B' + str(row)] = dicts[1]
        # ws['C' + str(row)] = dicts[2]
        # wb2.save('record.xlsx')

    print('共' + str(max_row-2) + '条下载链接。\n成功：' + str(count) + '\n失败：' + str(max_row -2 - count))
    end = datetime.now() - start
    print('用时：' + str(end))


if __name__ == '__main__':
    deal_excel(argv[1])
